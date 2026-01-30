from openpyxl import load_workbook
import pandas as pd
import degressieve_ul_llngroepen as dul
import ast
import re
import os
import subprocess
import get_data

# import flask/fastapi (test postman)

if 'ul_herwerking.xlsx' not in os.listdir('output'):
    subprocess.run(['python', 'scripts\\herwerking_ul_llngroepen.py'])

wb = load_workbook('Brondata\\simulatie_input.xlsx', data_only=True)
ws = wb['input']

sim_type = 'BEDRAG'

def get_llngroepen_alt(llngroepen, d):
    llngroepen = ast.literal_eval(re.sub(r'\bnan\b', 'None', llngroepen))
    result = {}
    for vp, llng in llngroepen.items():
        if pd.notna(llng):
            for key, value in llng.items():
                if key not in result:
                    result[key] = {'inschrijvingen': 0}
                result[key]['inschrijvingen'] += value['inschrijvingen']

    for key, value in result.items():
        result[key]['ul'] = dul.get_degressieve_uren_leraar(key, value['inschrijvingen'], True, d)
    return result

def ul_alt(llngroepen):
    result = 0
    for key, value in llngroepen.items():
        result += value['ul']
    return result

def punten_dir_alt(aantal, punten_per_lln):
    return aantal*punten_per_lln

def get_nieuwe_coef(groep, lln_tobe, ul):
    try:
        coef = dul.degressieve_ul[groep]['coef']
    except:
        return [0,0,0,0]

    lln_tobe = ast.literal_eval(lln_tobe)
    ul_herverdeeld = 0

    for i in range(len(coef) - 1, -1, -1):
        for j in range(i, len(coef)):
            coef[j] += (ul - ul_herverdeeld)/sum(lln_tobe[i:])
        if i==0 or coef[i] < coef[i-1]:
            break
        else:
            ul_herverdeeld = 0
            for j in range(i, len(coef)):
                coef[j] = coef[i-1]
                ul_herverdeeld += coef[j] * lln_tobe[j]
        
    return coef

def ul_punten_bedrag():
    df_ul_herverdeling = pd.read_excel('output\\ul_herwerking.xlsx')

    ul_besparing_units = ws['B4'].value
    ul_besparing_alt = ws['B7'].value
    print(f'BESPARING UREN-LERAAR: {ul_besparing_alt}')

    ul_herverdeling = ul_besparing_units - ul_besparing_alt
    df_ul_herverdeling['herverdeelde_ul'] = df_ul_herverdeling['herverdeling_percent']*ul_herverdeling
    df_ul_herverdeling['coef_alt'] = df_ul_herverdeling.apply(lambda row: get_nieuwe_coef(
        row['leerlingengroep'], row['lln_schijven_tobe'], row['herverdeelde_ul']
    ), axis=1)
    d = dict(zip(df_ul_herverdeling['leerlingengroep'],
                 df_ul_herverdeling['coef_alt']))

    punten_dir_asis = 992*120
    aantal_lln = 476901

    punten_besparing = ws['G7'].value
    print(f'BESPARING PUNTEN DIRECTEUR: {punten_besparing}')
    punten_dir_per_lln_besp = (punten_dir_asis - punten_besparing)/aantal_lln

    return d, punten_dir_per_lln_besp


if sim_type == 'BEDRAG':
    ul_dict, ptn_dir_lln = ul_punten_bedrag()
elif sim_type == 'COEF':
    ul_dict = {
        '1e graad A': ws['B12'].value,
        '1e graad B': ws['B13'].value,
        '2e graad bso': ws['B14'].value,
        '2e graad tso': ws['B15'].value,
        '3e graad bso': ws['B16'].value,
        '3e graad tso': ws['B17'].value,
        '2e graad kso': ws['B18'].value,
        '3e graad kso': ws['B19'].value,
        'okan': ws['B20'].value,
        'hbo': ws['B21'].value,
        '2e graad aso': ws['B22'].value,
        '3e graad aso': ws['B23'].value,
        'n.v.t. (modulair) bso': ws['B24'].value,
        '4e graad bso': ws['B25'].value,
    }
    ptn_dir_lln = ws['G13'].value


df_units = pd.read_excel('output\\jaren\\2024-2025\\8_analyse_units.xlsx')
df_units['llngr_ul_alt'] = df_units.apply(lambda row: get_llngroepen_alt(row['llng_asis'], ul_dict), axis=1)
df_units['ul_alt'] = df_units['llngr_ul_alt'].apply(ul_alt)
df_units['punten_dir_alt'] = df_units.apply(lambda row: punten_dir_alt(row['aantal_leerlingen'], ptn_dir_lln), axis=1)
df_units = df_units[df_units['aantal_leerlingen']>0]

ul_alt_tot = df_units['ul_vast'].sum() + df_units['ul_alt'].sum()
ptn_dir_alt_tot = df_units['punten_dir_alt'].sum()

ul_dict['punten_directeur_per_lln'] = ptn_dir_lln
ul_dict['lln_per_directeur'] = 120/ptn_dir_lln
ul_dict['punten_directuer_per_ul'] = ptn_dir_alt_tot/ul_alt_tot
ul_dict['uren-leraar_per_directeur'] = 120/ul_dict['punten_directuer_per_ul']

agg_cols = ['aantal_leerlingen', 'ul_vast', 'ul_asis', 'ul_alt', 'punten_dir_asis', 'punten_dir_alt']
group_cols = ['schoolbestuur', 'scholengemeenschap', 'net']

uitleg_var = {
    'aantal_leerlingen': 'Totaal aantal leerlingen',
    'ul_asis': 'Degressieve uren-leraar in de AS-IS (financiering op niveau Instelling)',
    'ul_alt': 'Degressieve uren-leraar in de TO-BE (financiering op niveau Units) gebruikmakend van de alternatieve regeling van de nieuwe coëfficiënt voor de laatste schijf leerlingen',
    'ul_diff': 'Verschil uren-leraar tussen AS-IS en TO-BE (ul_alt - ul_asis)',
    'ul_diff_euro': 'Verschil uren-leraar uitegedrukt in euro (ul_diff * 0.9657 * 69073 / 21.23)',
    'punten_dir_asis': 'Punten GBE die een scholengemeenschap zou krijgen voor directeurs in de AS-IS',
    'punten_dir_alt': 'Punten GBE voor directeurs in de TO-BE gebruikmakend van de alternatieve regeling met aantal leerlingen per directeur',
    'punten_dir_diff': 'Verschil punten GBE voor directeurs tussen AS-IS en TO-BE (punten_dir_alt - punten_dir_asis)',
    'punten_dir_diff_euro': 'Verschil punten GBE voor directeurs uitgedrukt in euro (punten_dir_diff * 752,4)',
    'diff_euro_totaal': 'Het totale verschil (ul_diff_euro + punten_dir_diff_euro)',
    'oki': 'Gewogen gemiddelde van de oki scores van de vestigingsplaatsen',
    'sheet coëfficiënten': 'De gebruikte coëfficiënten voor degressieve uren-leraar per leerlingengroep en het aantal studenten/uren-leraar per directeur (en afgeleide punten directeur per student/uren-leraar)'
}

with pd.ExcelWriter(f'output\\simulatie_alternatief_{sim_type}.xlsx') as writer:
    df_uitleg = pd.DataFrame(list(uitleg_var.items()), columns=['variabele', 'uitleg'])
    df_uitleg.to_excel(writer, sheet_name='uitleg variabelen', index=False)

    df_units['ul_diff'] = df_units['ul_alt'] - df_units['ul_asis']
    df_units['ul_diff_euro'] = df_units['ul_diff']*0.9657*69073/21.23
    df_units['punten_dir_diff'] = df_units['punten_dir_alt'] - df_units['punten_dir_asis']
    df_units['punten_dir_diff_euro'] = df_units['punten_dir_diff'] * 752.4
    df_units['diff_euro_totaal'] = df_units['ul_diff_euro'] + df_units['punten_dir_diff_euro']
    df_units['oki'] = df_units.apply(lambda row: get_data.get_oki(row['unit_code_so'], '2024-2025'), axis=1)
    df_units_xslx = df_units[['unit_code_so', 'aantal_leerlingen', 'ul_vast',
        'ul_asis', 'ul_alt', 'ul_diff', 'ul_diff_euro',
        'punten_dir_asis', 'punten_dir_alt', 'punten_dir_diff', 'punten_dir_diff_euro',
        'diff_euro_totaal', 'oki'
        ]]
    df_units_xslx.to_excel(writer, sheet_name='units', index=False)
    for groep in group_cols:
        df = df_units.groupby(groep)[agg_cols].sum().reset_index()
        df['ul_diff'] = df['ul_alt'] - df['ul_asis']
        df['ul_diff_euro'] = df['ul_diff']*0.9657*69073/21.23
        df['punten_dir_diff'] = df['punten_dir_alt'] - df['punten_dir_asis']
        df['punten_dir_diff_euro'] = df['punten_dir_diff'] * 752.4
        df['diff_euro_totaal'] = df['ul_diff_euro'] + df['punten_dir_diff_euro']
        df['oki'] = df_units.groupby(groep).apply(
            lambda row: get_data.get_oki('_'.join(row['unit_code_so']), '2024-2025'),
            include_groups=False
        ).values
        df = df[[groep, 'aantal_leerlingen', 'ul_vast', 'ul_asis', 'ul_alt', 'ul_diff', 'ul_diff_euro',
            'punten_dir_asis', 'punten_dir_alt', 'punten_dir_diff', 'punten_dir_diff_euro', 
            'diff_euro_totaal', 'oki'
            ]]
        df.to_excel(writer, sheet_name=groep, index=False)
        
    df_coef = pd.DataFrame(list(ul_dict.items()), columns=['coëfficiënt', 'waarde'])
    df_coef.to_excel(writer, sheet_name='coëfficiënten', index=False)


# request (teacher_hours, points_director)

# amount money: response: excel file, coefficients
#               coefficients: 
# {
#     '1e graad A': []
#     '1e graad B': []
#     ...
#     punten_dir: float
# }


# request (coefficients)

# coefficients: response: excel file, ul_besparing, dir_besparing, totaal_besparing